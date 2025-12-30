#!/usr/bin/env python3
"""Import MorphoLex-en morpheme data into tokenizer database"""

import re
import sqlite3
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl. Run: pip install openpyxl")
    exit(1)

# Paths - adjust these for your setup
SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent
DB_PATH = BASE_DIR / "db" / "tokenizer.db"
REF_PATH = BASE_DIR / "reference" / "morpholex"

SOURCE_NAME = "morpholex-en"
SOURCE_CONFIDENCE = 0.85

# Sheets to skip (presentation and summary sheets)
SKIP_SHEETS = {'Presentation', 'Prefixes', 'Suffixes', 'Roots'}

# Morpheme mappings - surface form to concept
PREFIX_TO_CONCEPT = {
    'a': 'NEGATION',  # as in 'amoral'
    'un': 'NEGATION',
    'in': 'NEGATION',
    'im': 'NEGATION',
    'il': 'NEGATION',
    'ir': 'NEGATION',
    'dis': 'NEGATION',
    'non': 'NEGATION',
    'de': 'REVERSAL',
    'anti': 'AGAINST',
    're': 'AGAIN',
    'pre': 'BEFORE',
    'post': 'AFTER',
    'sub': 'UNDER',
    'super': 'ABOVE',
    'over': 'ABOVE',
    'under': 'UNDER',
    'inter': 'BETWEEN',
    'trans': 'ACROSS',
    'ex': 'OUT',
    'out': 'OUT',
    'en': 'CAUSATIVE',
    'em': 'CAUSATIVE',
    'co': 'TOGETHER',
    'com': 'TOGETHER',
    'con': 'TOGETHER',
    'multi': 'MANY',
    'poly': 'MANY',
    'mono': 'ONE',
    'uni': 'ONE',
    'bi': 'TWO',
    'di': 'TWO',
    'tri': 'THREE',
    'semi': 'HALF',
    'auto': 'SELF',
    'self': 'SELF',
    'mis': 'WRONG',
    'mal': 'BAD',
    'bene': 'GOOD',
}

SUFFIX_TO_CONCEPT = {
    'er': 'AGENT_OF',
    'or': 'AGENT_OF',
    'ist': 'AGENT_OF',
    'ian': 'AGENT_OF',
    'ant': 'AGENT_OF',
    'ent': 'AGENT_OF',
    'ee': 'PATIENT_OF',
    'ness': 'STATE_OF',
    'ment': 'STATE_OF',
    'tion': 'STATE_OF',
    'sion': 'STATE_OF',
    'ity': 'STATE_OF',
    'ty': 'STATE_OF',
    'ance': 'STATE_OF',
    'ence': 'STATE_OF',
    'dom': 'STATE_OF',
    'hood': 'STATE_OF',
    'ship': 'STATE_OF',
    'ful': 'FULL_OF',
    'ous': 'QUALITY_OF',
    'ious': 'QUALITY_OF',
    'eous': 'QUALITY_OF',
    'al': 'QUALITY_OF',
    'ial': 'QUALITY_OF',
    'ic': 'QUALITY_OF',
    'ical': 'QUALITY_OF',
    'ive': 'QUALITY_OF',
    'less': 'WITHOUT',
    'able': 'CAPABLE_OF',
    'ible': 'CAPABLE_OF',
    'ly': 'MANNER',
    'ize': 'CAUSATIVE',
    'ise': 'CAUSATIVE',
    'ify': 'CAUSATIVE',
    'en': 'CAUSATIVE',
    'ate': 'CAUSATIVE',
    'ed': 'PAST',
    'ing': 'CONTINUOUS',
    's': 'PLURAL',
    'es': 'PLURAL',
    'ward': 'DIRECTION',
    'wards': 'DIRECTION',
    'wise': 'MANNER',
    'like': 'SIMILAR',
    'ish': 'SOMEWHAT',
    'most': 'SUPERLATIVE',
    'est': 'SUPERLATIVE',
}


def get_or_create_source(cur) -> int:
    """Register this source in data_sources table"""
    cur.execute("""
        INSERT OR IGNORE INTO data_sources (name, source_type, base_confidence)
        VALUES (?, ?, ?)
    """, (SOURCE_NAME, 'corpus', SOURCE_CONFIDENCE))
    cur.execute("SELECT id FROM data_sources WHERE name = ?", (SOURCE_NAME,))
    return cur.fetchone()[0]


def get_english_language_id(cur) -> int:
    """Get or create English language entry"""
    cur.execute("SELECT id FROM language_families WHERE iso_639_3 = 'eng'")
    result = cur.fetchone()
    if result:
        return result[0]

    cur.execute("""
        INSERT INTO language_families (name, level, iso_639_3, source)
        VALUES ('English', 'language', 'eng', 'morpholex')
    """)
    return cur.lastrowid


def find_column_indices(headers):
    """Find column indices for the fields we need"""
    indices = {'word': None, 'pos': None, 'segm': None}

    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str == 'Word':
            indices['word'] = i
        elif h_str == 'POS':
            indices['pos'] = i
        elif h_str == 'MorphoLexSegm':
            indices['segm'] = i

    return indices


def parse_segmentation(segm_str):
    """
    Parse MorphoLex segmentation notation.
    Example: {<a<(litera)>ate>} -> prefix 'a', root 'litera', suffix 'ate'

    Notation:
    - (...) = root
    - <...> around ( = prefix before root
    - <...> after ) = suffix after root
    - {...} = word boundary
    """
    if not segm_str:
        return {'prefixes': [], 'roots': [], 'suffixes': []}

    prefixes = []
    roots = []
    suffixes = []

    # Extract roots first - they're in parentheses
    root_matches = re.findall(r'\(([^)]+)\)', segm_str)
    roots = [r.strip() for r in root_matches if r.strip()]

    # Remove the root portions to analyze affixes
    # Pattern: stuff before ( is prefixes, stuff after ) is suffixes
    # Work through the structure

    # Simple approach: split by roots and analyze
    # For {<a<(litera)>ate>}:
    # - Before first ( : {<a<  -> extract 'a'
    # - After last ) : >ate>} -> extract 'ate'

    # Find content before first root marker
    first_paren = segm_str.find('(')
    if first_paren > 0:
        prefix_part = segm_str[:first_paren]
        # Extract alphabetic sequences as prefixes
        prefix_matches = re.findall(r'[a-zA-Z]+', prefix_part)
        prefixes = [p.lower() for p in prefix_matches if p]

    # Find content after last root marker
    last_paren = segm_str.rfind(')')
    if last_paren >= 0 and last_paren < len(segm_str) - 1:
        suffix_part = segm_str[last_paren + 1:]
        # Extract alphabetic sequences as suffixes
        suffix_matches = re.findall(r'[a-zA-Z]+', suffix_part)
        suffixes = [s.lower() for s in suffix_matches if s]

    return {
        'prefixes': prefixes,
        'roots': roots,
        'suffixes': suffixes
    }


def parse_morpholex(ref_path: Path):
    """Parse all data sheets from MorphoLex Excel file"""
    xlsx_file = None
    for f in ref_path.glob("*.xlsx"):
        xlsx_file = f
        break

    if not xlsx_file:
        raise FileNotFoundError(f"No Excel file found in {ref_path}")

    print(f"  Reading {xlsx_file.name}...")

    # Don't use read_only mode - it has issues with this file
    wb = openpyxl.load_workbook(xlsx_file, read_only=False, data_only=True)

    sheet_count = 0
    row_count = 0

    for sheet_name in wb.sheetnames:
        # Skip non-data sheets
        if sheet_name in SKIP_SHEETS:
            continue

        # Data sheets are named like '0-1-0', '1-1-1', etc.
        if not re.match(r'^\d+-\d+-\d+$', sheet_name):
            continue

        sheet_count += 1
        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        indices = find_column_indices(headers)

        if indices['word'] is None:
            print(f"    Skipping sheet {sheet_name} - no Word column found")
            continue

        # Process data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= indices['word']:
                continue

            word = row[indices['word']]
            if not word:
                continue

            pos = row[indices['pos']] if indices['pos'] is not None and indices['pos'] < len(row) else None
            segm = row[indices['segm']] if indices['segm'] is not None and indices['segm'] < len(row) else None

            row_count += 1

            # Parse segmentation
            morphemes = parse_segmentation(segm) if segm else {'prefixes': [], 'roots': [], 'suffixes': []}

            yield {
                'word': str(word).strip().lower(),
                'pos': str(pos).strip() if pos else '',
                'segmentation': str(segm) if segm else '',
                'prefixes': morphemes['prefixes'],
                'roots': morphemes['roots'],
                'suffixes': morphemes['suffixes'],
                'prs_signature': sheet_name,
            }

    wb.close()
    print(f"  Processed {sheet_count} sheets, {row_count} words")


def import_morpholex(conn, ref_path: Path, source_id: int):
    """Import MorphoLex data"""
    cur = conn.cursor()

    eng_id = get_english_language_id(cur)
    print(f"  English language ID: {eng_id}")

    # Build morpheme concept lookup
    cur.execute("SELECT id, morpheme FROM morphemes")
    concept_to_id = {row[1]: row[0] for row in cur.fetchall()}
    print(f"  Found {len(concept_to_id)} morpheme concepts")

    insert_count = 0
    link_count = 0
    batch = []
    batch_size = 500

    for rec in parse_morpholex(ref_path):
        # Determine primary morpheme_id from affixes
        morpheme_id = None

        # Check prefixes for concept match
        for prefix in rec['prefixes']:
            concept = PREFIX_TO_CONCEPT.get(prefix)
            if concept and concept in concept_to_id:
                morpheme_id = concept_to_id[concept]
                link_count += 1
                break

        # Check suffixes for concept match
        if not morpheme_id:
            for suffix in rec['suffixes']:
                concept = SUFFIX_TO_CONCEPT.get(suffix)
                if concept and concept in concept_to_id:
                    morpheme_id = concept_to_id[concept]
                    link_count += 1
                    break

        batch.append((
            morpheme_id,
            eng_id,
            rec['word'],
            0,  # frequency - not in per-word data
            rec['pos'],
            rec['segmentation'],
            SOURCE_NAME,
        ))

        if len(batch) >= batch_size:
            cur.executemany("""
                INSERT OR IGNORE INTO surface_forms
                (morpheme_id, language_id, surface_form, frequency, pos_tag, pos_features, source)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, batch)
            insert_count += cur.rowcount
            conn.commit()
            batch = []

            if insert_count % 10000 == 0:
                print(f"    Inserted {insert_count} surface forms...")

    # Final batch
    if batch:
        cur.executemany("""
            INSERT OR IGNORE INTO surface_forms
            (morpheme_id, language_id, surface_form, frequency, pos_tag, pos_features, source)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, batch)
        insert_count += cur.rowcount
        conn.commit()

    print(f"  Inserted {insert_count} surface forms")
    print(f"  Linked {link_count} to morpheme concepts")


def main():
    print(f"Importing MorphoLex-en from {REF_PATH}")

    if not REF_PATH.exists():
        print(f"ERROR: Reference directory not found: {REF_PATH}")
        print("Run download_sources.py first")
        return 1

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    try:
        cur = conn.cursor()
        source_id = get_or_create_source(cur)
        conn.commit()

        import_morpholex(conn, REF_PATH, source_id)

        # Validation
        cur.execute("""
            SELECT COUNT(*) FROM surface_forms WHERE source = ?
        """, (SOURCE_NAME,))
        count = cur.fetchone()[0]
        print(f"\n  Total surface forms from MorphoLex: {count}")

        cur.execute("""
            SELECT COUNT(*) FROM surface_forms
            WHERE source = ? AND morpheme_id IS NOT NULL
        """, (SOURCE_NAME,))
        linked = cur.fetchone()[0]
        print(f"  Linked to concepts: {linked}")

    finally:
        conn.close()

    print("\nMorphoLex import complete!")
    return 0


if __name__ == '__main__':
    exit(main())
